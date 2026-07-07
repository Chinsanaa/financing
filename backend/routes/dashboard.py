"""Dashboard endpoints: stats, trends, category breakdowns.

Notes on postgrest-py usage: negated filters use the `.not_` PROPERTY
(e.g. `.not_.is_("category_id", "null")`). Calling `.not_(...)` raises
TypeError — it is not a method.
"""
from fastapi import APIRouter, HTTPException, Request
from pydantic import BaseModel
from typing import List, Optional
from config import supabase_client
from db import fetch_all
from errors import internal_error
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
import pandas as pd
from translate import merchant_label_english, description_label_english

router = APIRouter()


def _now_cn() -> datetime:
    """Naive China-clock 'now'. Transaction timestamps come from Alipay/WeChat
    exports in China local time and are stored naive, so month boundaries and
    cutoffs must use the same clock — not the server's (UTC on Railway)."""
    return datetime.now(ZoneInfo("Asia/Shanghai")).replace(tzinfo=None)


def _month_start(now: datetime = None) -> datetime:
    now = now or _now_cn()
    return datetime(now.year, now.month, 1)


@router.get("/summary")
async def get_summary(request: Request):
    """Overall summary: total transactions, labeled%, total spend."""
    user_id = request.state.user_id

    try:
        # Total transactions
        total = supabase_client.table("transactions").select("id", count="exact").eq("user_id", user_id).execute()
        total_count = total.count if total.count is not None else len(total.data)

        # Labeled = has a trusted category (not still pending review)
        labeled = (
            supabase_client.table("transactions")
            .select("id", count="exact")
            .eq("user_id", user_id)
            .eq("needs_review", False)
            .not_.is_("category_id", "null")
            .execute()
        )
        labeled_count = labeled.count if labeled.count is not None else len(labeled.data)

        # Total spend (fetch_all pages past PostgREST's 1000-row cap, which
        # silently truncated — and understated — every sum here before)
        all_transactions = fetch_all(
            lambda: supabase_client.table("transactions").select("amount").eq("user_id", user_id)
        )
        total_spend = sum(float(t["amount"]) for t in all_transactions)

        return {
            "total_transactions": total_count,
            "labeled_transactions": labeled_count,
            "labeling_percentage": round(100 * labeled_count / total_count, 1) if total_count > 0 else 0,
            "total_spend": total_spend,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise internal_error(e, "dashboard/summary")


@router.get("/by-category")
async def get_by_category(request: Request):
    """Spending breakdown by category (transactions with category_id set)."""
    user_id = request.state.user_id

    try:
        rows = fetch_all(
            lambda: supabase_client.table("transactions")
            .select("amount, category_id, categories(name)")
            .eq("user_id", user_id)
            .not_.is_("category_id", "null")
        )

        if not rows:
            return {"categories": []}

        df = pd.DataFrame(rows)
        df["category_name"] = df["categories"].apply(lambda x: x["name"] if x else "Unknown")

        breakdown = df.groupby("category_name")["amount"].agg(["sum", "count"]).reset_index()
        breakdown.columns = ["category", "total_amount", "count"]

        return {
            "categories": [
                {
                    "category": row["category"],
                    "total_amount": float(row["total_amount"]),
                    "transaction_count": int(row["count"]),
                }
                for _, row in breakdown.iterrows()
            ]
        }
    except HTTPException:
        raise
    except Exception as e:
        raise internal_error(e, "dashboard/by-category")


@router.get("/trends")
async def get_trends(request: Request, days: int = 30):
    """Spending trends: daily spend over last N days."""
    user_id = request.state.user_id

    try:
        cutoff = _now_cn() - timedelta(days=days)
        # Filter server-side; never pull the full transaction history.
        rows = fetch_all(
            lambda: supabase_client.table("transactions")
            .select("timestamp, amount")
            .eq("user_id", user_id)
            .not_.is_("category_id", "null")
            .gte("timestamp", cutoff.isoformat())
        )

        if not rows:
            return {"trends": []}

        df = pd.DataFrame(rows)
        df["date"] = pd.to_datetime(df["timestamp"]).dt.date
        daily = df.groupby("date")["amount"].sum().reset_index()

        return {
            "trends": [
                {
                    "date": str(row["date"]),
                    "total_spend": float(row["amount"]),
                }
                for _, row in daily.iterrows()
            ]
        }
    except HTTPException:
        raise
    except Exception as e:
        raise internal_error(e, "dashboard/trends")


def _monthly_income(user_id: str) -> float:
    """Monthly income from profiles.monthly_income (the only written source)."""
    resp = supabase_client.table("profiles").select("monthly_income").eq("id", user_id).execute()
    if resp.data and resp.data[0].get("monthly_income") is not None:
        return float(resp.data[0]["monthly_income"])
    return 0.0


def _budget_config(user_id: str):
    resp = supabase_client.table("budget_config").select("*").eq("user_id", user_id).execute()
    return resp.data[0] if resp.data else None


def _current_month_spend_by_category(user_id: str) -> dict:
    """This month's spend per category name (monthly budgets compare against
    the current month, not all-time history)."""
    rows = fetch_all(
        lambda: supabase_client.table("transactions")
        .select("amount, category_id, categories(name)")
        .eq("user_id", user_id)
        .not_.is_("category_id", "null")
        .gte("timestamp", _month_start().isoformat())
    )
    if not rows:
        return {}
    df = pd.DataFrame(rows)
    df["category_name"] = df["categories"].apply(lambda x: x["name"] if x else "Unknown")
    return df.groupby("category_name")["amount"].sum().to_dict()


@router.get("/budget")
async def get_budget(request: Request):
    """Budget info: limits by category, current-month spend vs budget."""
    user_id = request.state.user_id

    try:
        # Monthly income's single source of truth is profiles.monthly_income
        # (written by PATCH /settings/profile). budget_config.income is legacy
        # and was never written by anything.
        budget_config = _budget_config(user_id)

        cat_budget_resp = (
            supabase_client.table("budget_category_config")
            .select("*, categories(name)")
            .eq("user_id", user_id)
            .execute()
        )

        budget_config_out = {
            "monthly_income": _monthly_income(user_id),
            "currency": budget_config["currency"] if budget_config else "CNY",
        }

        if not cat_budget_resp.data:
            return {"budget_config": budget_config_out, "category_budgets": []}

        spending_by_cat = _current_month_spend_by_category(user_id)

        category_budgets = []
        for row in cat_budget_resp.data:
            cat_name = row["categories"]["name"] if row["categories"] else "Unknown"
            category_budgets.append({
                "category": cat_name,
                "monthly_budget": float(row["monthly_budget"]) if row["monthly_budget"] else 0,
                "current_spend": float(spending_by_cat.get(cat_name, 0)),
                "type": row["type"],
            })

        return {
            "budget_config": budget_config_out,
            "category_budgets": category_budgets,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise internal_error(e, "dashboard/budget")


class CategoryBudgetItem(BaseModel):
    category_id: str
    monthly_budget: Optional[float] = None
    type: str = "Need"  # budget_type enum: 'Need' | 'Want'


class CategoryBudgetsUpdate(BaseModel):
    budgets: List[CategoryBudgetItem]


@router.put("/budget/categories")
async def put_category_budgets(request: Request, data: CategoryBudgetsUpdate):
    """Set per-category monthly budgets (upsert on user_id + category_id).

    This is the writer budget_category_config never had — the Budget tab and
    the over-budget Action items read from it but nothing could populate it.
    """
    user_id = request.state.user_id

    try:
        if not data.budgets:
            raise HTTPException(status_code=400, detail="No budgets provided")

        # Only allow the user's own categories
        cat_resp = supabase_client.table("categories").select("id").eq("user_id", user_id).execute()
        own_ids = {c["id"] for c in (cat_resp.data or [])}

        rows = []
        for item in data.budgets:
            if item.category_id not in own_ids:
                raise HTTPException(status_code=400, detail="Unknown category")
            if item.type not in ("Need", "Want"):
                raise HTTPException(status_code=400, detail="type must be 'Need' or 'Want'")
            rows.append({
                "user_id": user_id,
                "category_id": item.category_id,
                "type": item.type,
                "monthly_budget": item.monthly_budget,
            })

        response = (
            supabase_client.table("budget_category_config")
            .upsert(rows, on_conflict="user_id,category_id")
            .execute()
        )
        return {"updated": len(response.data or rows)}
    except HTTPException:
        raise
    except Exception as e:
        raise internal_error(e, "dashboard/put_category_budgets")


@router.get("/savings")
async def get_savings(request: Request):
    """Savings info: goals, current savings, anomalies."""
    user_id = request.state.user_id

    try:
        budget_config = _budget_config(user_id)

        now = _now_cn()
        month_start = _month_start(now)

        spending_rows = fetch_all(
            lambda: supabase_client.table("transactions")
            .select("amount")
            .eq("user_id", user_id)
            .gte("timestamp", month_start.isoformat())
        )
        current_spend = sum(float(t["amount"]) for t in spending_rows)

        # Simple anomaly detection: compare to average of last 3 months
        three_months_ago = datetime(now.year if now.month >= 4 else now.year - 1,
                                    now.month - 3 if now.month >= 4 else now.month + 9, 1)

        historical_rows = fetch_all(
            lambda: supabase_client.table("transactions")
            .select("timestamp, amount")
            .eq("user_id", user_id)
            .gte("timestamp", three_months_ago.isoformat())
            .lt("timestamp", month_start.isoformat())
        )

        historical_df = pd.DataFrame(historical_rows) if historical_rows else pd.DataFrame()
        avg_monthly = 0
        if not historical_df.empty:
            historical_df["month"] = pd.to_datetime(historical_df["timestamp"]).dt.to_period("M")
            monthly_totals = historical_df.groupby("month")["amount"].sum()
            avg_monthly = float(monthly_totals.mean()) if len(monthly_totals) > 0 else 0

        savings_goal = float(budget_config["saving_goal_monthly"]) if budget_config and budget_config["saving_goal_monthly"] else 0
        income = _monthly_income(user_id)

        return {
            "savings_goal_monthly": savings_goal,
            "income": income,
            "current_spend": float(current_spend),
            "projected_savings": max(0, income - current_spend),
            "average_monthly_spend": avg_monthly,
            "is_anomaly": current_spend > (avg_monthly * 1.3) if avg_monthly > 0 else False,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise internal_error(e, "dashboard/savings")


@router.get("/action")
async def get_action(request: Request):
    """Action items: over-budget categories, review queue count."""
    user_id = request.state.user_id

    try:
        actions = []

        cat_budget_resp = (
            supabase_client.table("budget_category_config")
            .select("*, categories(name)")
            .eq("user_id", user_id)
            .execute()
        )

        if cat_budget_resp.data:
            spending_by_cat = _current_month_spend_by_category(user_id)

            for budget_row in cat_budget_resp.data:
                cat_name = budget_row["categories"]["name"] if budget_row["categories"] else "Unknown"
                budget = float(budget_row["monthly_budget"]) if budget_row["monthly_budget"] else 0
                spend = float(spending_by_cat.get(cat_name, 0))

                if budget > 0 and spend > budget:
                    actions.append({
                        "type": "over_budget",
                        "category": cat_name,
                        "current": spend,
                        "limit": budget,
                        "overage": spend - budget,
                    })

        # Review queue count
        review_resp = (
            supabase_client.table("transactions")
            .select("id", count="exact")
            .eq("user_id", user_id)
            .eq("needs_review", True)
            .execute()
        )
        review_count = review_resp.count if review_resp.count is not None else len(review_resp.data)
        if review_count > 0:
            actions.append({
                "type": "pending_review",
                "count": review_count,
                "message": f"{review_count} transactions need review",
            })

        return {"actions": actions}
    except HTTPException:
        raise
    except Exception as e:
        raise internal_error(e, "dashboard/action")


@router.get("/reports")
async def get_reports(request: Request, page: int = 1, per_page: int = 100):
    """Detailed reports: paginated categorized-transaction list."""
    user_id = request.state.user_id
    page = max(1, page)
    per_page = min(max(1, per_page), 500)

    try:
        start = (page - 1) * per_page
        response = (
            supabase_client.table("transactions")
            .select("timestamp, merchant, description, amount, category_id, categories(name), label_source",
                    count="exact")
            .eq("user_id", user_id)
            .not_.is_("category_id", "null")
            .order("timestamp", desc=True)
            .range(start, start + per_page - 1)
            .execute()
        )

        total_count = response.count if response.count is not None else len(response.data or [])
        transactions = [
            {
                "date": txn["timestamp"],
                "merchant": merchant_label_english(txn["merchant"]),
                "description": description_label_english(txn["description"]),
                "amount": float(txn["amount"]),
                "category": txn["categories"]["name"] if txn["categories"] else "Unknown",
                "label_source": txn["label_source"],
            }
            for txn in (response.data or [])
        ]

        return {
            "transactions": transactions,
            "total_count": total_count,
            "page": page,
            "per_page": per_page,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise internal_error(e, "dashboard/reports")


@router.get("/export")
async def export_transactions(request: Request):
    """Export all user transactions as XLSX workbook.

    Columns: Date | Merchant | Description | Category | Amount | Label source
    All text is translated to English; amounts formatted as #,##0.00
    """
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import datetime

    user_id = request.state.user_id

    try:
        def make_query():
            return (
                supabase_client.table("transactions")
                .select("timestamp, merchant, description, amount, category_id, categories(name), label_source")
                .eq("user_id", user_id)
                .order("timestamp", desc=True)
            )

        all_txns = fetch_all(make_query)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"

        # Headers
        headers = ["Date", "Merchant", "Description", "Category", "Amount", "Label Source"]
        ws.append(headers)

        # Format header row
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Freeze header
        ws.freeze_panes = "A2"

        # Add data rows
        for txn in all_txns:
            ws.append([
                txn["timestamp"],
                merchant_label_english(txn["merchant"]),
                description_label_english(txn["description"]),
                txn["categories"]["name"] if txn["categories"] else "Uncategorized",
                float(txn["amount"]),
                txn["label_source"] or "",
            ])

        # Format amount column
        for row in ws.iter_rows(min_row=2, max_row=len(all_txns) + 1, min_col=5, max_col=5):
            for cell in row:
                cell.number_format = '#,##0.00'

        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 14

        # Write to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Return as attachment
        now = datetime.now().strftime("%Y%m%d_%H%M%S")
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=transactions_{now}.xlsx"}
        )
    except Exception as e:
        raise internal_error(e, "dashboard/export")


@router.get("/review-queue")
async def get_review_queue(request: Request, show_labeled: bool = False):
    """Transactions needing review.

    By default (show_labeled=False): model suggestions with needs_review=True,
    sorted by confidence (lowest first — least confident suggestions).

    When show_labeled=True: manually labeled transactions (is_manually_labeled=True),
    so user can audit their own labels and catch human errors. Useful for catching
    accidental miscategorizations before retraining.
    """
    user_id = request.state.user_id

    try:
        from src.translate import merchant_label_english, description_label_english

        if show_labeled:
            # Show manually labeled transactions for user review/correction
            response = (
                supabase_client.table("transactions")
                .select("id, timestamp, merchant, description, amount, confidence, category_id, categories(name)")
                .eq("user_id", user_id)
                .eq("is_manually_labeled", True)
                .order("timestamp", desc=True)  # Most recent labels first
                .limit(50)
                .execute()
            )
            label_type = "manually_labeled"
        else:
            # Show model suggestions needing review
            response = (
                supabase_client.table("transactions")
                .select("id, timestamp, merchant, description, amount, confidence, category_id, categories(name)")
                .eq("user_id", user_id)
                .eq("needs_review", True)
                .order("confidence")  # Least confident first
                .limit(50)
                .execute()
            )
            label_type = "model_suggestion"

        if not response.data:
            return {"transactions": [], "count": 0, "type": label_type}

        # In suggestion mode the row's category_id IS the model's suggestion —
        # report it only as suggested_category so the UI can distinguish
        # "this is what it's labeled" from "this is what the model proposes".
        transactions = [
            {
                "id": txn["id"],
                "date": txn["timestamp"],
                "merchant": merchant_label_english(txn["merchant"]),
                "description": description_label_english(txn["description"]),
                "amount": float(txn["amount"]),
                "confidence": float(txn["confidence"]) if txn["confidence"] else 0,
                "category": (txn["categories"]["name"] if txn["categories"] else None) if show_labeled else None,
                "suggested_category": txn["categories"]["name"] if txn["categories"] else None,
            }
            for txn in response.data
        ]

        return {
            "transactions": transactions,
            "count": len(transactions),
            "type": label_type,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise internal_error(e, "dashboard/review-queue")


@router.get("/onboarding-status")
async def get_onboarding_status(request: Request):
    """Check user's onboarding progress."""
    user_id = request.state.user_id

    try:
        # profiles.id IS the auth user id (PK referencing auth.users)
        response = supabase_client.table("profiles").select("onboarding_phase").eq("id", user_id).execute()
        if not response.data:
            # 'upload' is the enum's first phase; 'signup' is not a valid value
            return {"onboarding_phase": "upload"}

        return {"onboarding_phase": response.data[0]["onboarding_phase"]}
    except HTTPException:
        raise
    except Exception as e:
        raise internal_error(e, "dashboard/onboarding-status")


@router.post("/onboarding-complete")
async def complete_onboarding(request: Request):
    """Mark onboarding as complete."""
    user_id = request.state.user_id

    try:
        supabase_client.table("profiles").update({
            "onboarding_phase": "complete"
        }).eq("id", user_id).execute()

        return {"message": "Onboarding complete"}
    except HTTPException:
        raise
    except Exception as e:
        raise internal_error(e, "dashboard/onboarding-complete")
